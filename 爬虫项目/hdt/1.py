import requests
from lxml import etree
from openpyxl import load_workbook
from time import sleep

headers2 = { 'referer': None, 'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36'}
headers1 = { 'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36'}

def GetHTMLContent(url): #return r
	headers = {'User-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36'}
	r = requests.get(url, headers = headers)
	r.raise_for_status()
	r.encoding = 'UTF-8'
	return r

def SaveList(url):
	r = GetHTMLContent(url) #得到总网页
	html = etree.HTML(r.text)
	T = html.xpath('/html/body/div[2]/div[1]/div[2]/ul/li/p[2]/a/@href') #得到所有页面的地址
	wb = load_workbook('data.xlsx')
	ws = wb['Sheet1']
	n=1
	for i in T:
		ws['A'+str(n+5558)] = T[n-1]
		n=n+1
	wb.save('data.xlsx')

def GetPic(url):
	r = GetHTMLContent(url) #得到第一页的页面
	html = etree.HTML(r.text)
	n = html.xpath('/html/body/div[2]/div[1]/div[4]/a[last()-1]/span')[0].text #得到页码
	# print(n)
	name = html.xpath('/html/body/div[2]/div[1]/h2')[0].text #得到图集名字
	# print(name)
	T = html.xpath('/html/body/div[2]/div[1]/div[3]/p/a/img/@src')[0][:-6] #得到
	print(T)
	return int(n), name, T

def SaveInfos():
	wb = load_workbook('data.xlsx')
	ws = wb['Sheet1']
	n = 800
	for i in ws['A'][n:]:
		url = i.value
		print(url)
		try:
			a, b, c = GetPic(url)
			# sleep(1)
			ws['B'+str(n)] = a
			ws['C'+str(n)] = b
			ws['D'+str(n)] = c
			print(n)
		except:
		n = n + 1
		sleep(2)
		if n == 1000:
			break
	wb.save('data.xlsx')

SaveInfos()