import requests
from lxml import etree
from openpyxl import Workbook, load_workbook
import sys
from tqdm import tqdm

def GetHTMLContent(url): #return r
    headers = {'User-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36'}
    r = requests.get(url, headers = headers)
    r.raise_for_status()
    r.encoding = 'UTF-8'
    return r

def get_last_page():
	url = 'https://wnacg.org/albums-index-page-1-cate-9.html'
	r = GetHTMLContent(url)
	html = etree.HTML(r.text)
	last_page = html.xpath('//*[@id="bodywrap"]/div[2]/div[2]/div/a[5]')[0].text
	return int(last_page)

def get_this_page(url):
	r = GetHTMLContent(url)
	html = etree.HTML(r.text)
	Ls = html.xpath('//*[@id="bodywrap"]/div[2]/div[1]/ul/li/div[1]/a/@href')
	return Ls

def form_url(last_page):
	url_vor = 'https://wnacg.org/albums-index-page-'
	url_nach = '-cate-9.html'
	urls = []
	for num in range(1, last_page + 1):
		url = url_vor + str(num) + url_nach
		urls.append(url)
	return urls

def enter_the_download_pages():
	last_page = get_last_page()
	urls = form_url(last_page)
	download_urls = []
	for url in tqdm(urls):
		L = get_this_page(url)
		for l in L:
			idx = l.split('.')[0].split('-')[-1]
			download_urls.append('https://wnacg.org/download-index-aid-' + idx + '.html')
	wb = load_workbook('data.xlsx')
	ws = wb['sheet1']
	for i in tqdm(range(1, len(download_urls) + 1)):
		ws['A' + str(i)] = download_urls[i - 1]
		if i % 50 == 0:
			wb.save('data.xlsx')
	wb.save('data.xlsx')

def zip_url():
	wb = load_workbook('data.xlsx')
	ws = wb['sheet1']
	for i in tqdm(range(1, 1001)):
		url = ws['A' + str(i)].value
		html = etree.HTML(GetHTMLContent(url).text)
		butten = html.xpath('/html/body/div[2]/div[2]/div/a[1]/@href')[0]
		ws['B' + str(i)] = butten
		if i % 50 == 0:
			wb.save('data.xlsx')
	wb.save('data.xlsx')
	wb.close()

def download_zip():
	wb = load_workbook('data.xlsx')
	ws = wb['sheet1']
	for i in tqdm(range(21, 31)):
		url = ws['B' + str(i)].value
		try:
			print('正在下载文件')
			r = GetHTMLContent(url).content
			size = sys.getsizeof(r) / 1024**2
			print(f'该文件大小为{size}Mb')
			with open(str(i) + '.zip', 'wb') as f:
				f.write(r)
			r = []
		except:
			ws['C' + str(i)] = 'error'
			print('获取文件失败，已记录并跳转至下一个文件')
			continue


if __name__ == '__main__':
	download_zip()