from selenium import webdriver
from lxml import etree
from time import sleep
from openpyxl import Workbook, load_workbook
import requests
import os


def GetHTMLContent(url):  # return r
	headers = {
		'User-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36'}
	r = requests.get(url, headers=headers)
	r.raise_for_status()
	r.encoding = 'UTF-8'
	return r


def get_the_url():
	driver = webdriver.Chrome()
	driver.get('https://nshens.com/?s=%E5%88%98%E9%A3%9E%E5%84%BF')
	element = driver.find_element_by_xpath('//*[@id="navigation-ajax"]')
	print('开始点击下一页')
	try:
		while 1:
			element.click()
			sleep(5)
	except:
		pass
	print('点击完毕')
	source = driver.page_source
	html = etree.HTML(source)
	titles = html.xpath(
		'//*[@id="cactus-body-container"]/div/div/div/div[2]/div/div[5]/div/div[1]/article/div/div[1]/div/a/@title')
	hrefs = html.xpath(
		'//*[@id="cactus-body-container"]/div/div/div/div[2]/div/div[5]/div/div[1]/article/div/div[1]/div/a/@href')
	print('已取得所有链接，开始写入数据库')
	wb = load_workbook('data.xlsx')
	ws = wb['sheet3']
	print('数据库已打开')
	for i in range(min(len(titles), len(hrefs))):
		ws['A' + str(i + 1)] = titles[i]
		ws['B' + str(i + 1)] = hrefs[i]
	print('写入完成')
	wb.save('data.xlsx')
	print('保存完成，程序结束')
	driver.close()


def get_the_number():
	wb = load_workbook('data.xlsx')
	ws = wb['sheet3']
	for i in range(1, 36):
		url = ws['B' + str(i)].value
		r = GetHTMLContent(url)
		sleep(1)
		html = etree.HTML(r.text)
		num = len(html.xpath(
			'//*[@id="cactus-body-container"]/div/div/div/div[2]/div/div/article/div[4]/div[3]/ul/li'))
		ws['C' + str(i)] = num
		print(i, num)
	wb.save('data.xlsx')


def pic_url():
	wb = load_workbook('data.xlsx')
	ws = wb['sheet3']
	for i in range(1, 36):
		url = ws['B' + str(i)].value
		# num = ws['C' + str(i)].value
		driver = webdriver.Chrome()
		driver.get(url)
		r = driver.page_source
		html = etree.HTML(r)
		try:
			erst_href = html.xpath(
				'//*[@class="seriesContent_li"][1]/div/a/@href')
			ws['D' + str(i)] = erst_href[0]
		except:
			pass
		driver.close()
		print(i)
		wb.save('data.xlsx')


def unterladen_pic():
	wb = load_workbook('data.xlsx')
	ws = wb['sheet3']
	for i in range(1, 36):
		name = ws['A' + str(i)].value
		num = int(name[-3:-1])
		url_root = ws['D' + str(i)].value
		url_root = url_root[:-5]
		# url_root = url_root[:-6]
		if not(os.path.exists(name)):
			os.mkdir(name)
		for j in range(1, num):
			# if j < 10:
			# 	j = '0' + str(j)
			# else:
			j = str(j)
			print(url_root + j + '.jpg')
			pic = GetHTMLContent(url_root + j + '.jpg').content
			with open(name + '/' + j + '.jpg', 'wb') as f:
				f.write(pic)


unterladen_pic()