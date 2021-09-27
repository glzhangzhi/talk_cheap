import requests
from lxml import etree
import time
import os
from openpyxl import load_workbook 

def GetHTMLContent(url): 
	headers = {'User-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:60.0) Gecko/20100101 Firefox/60.0'}
	r = requests.get(url, headers = headers)
	r.raise_for_status()
	r.encoding = 'UTF-8'
	return r

def main():
	for j in range(1,83):

		url = 'https://javzoo.com/cn/star/2jv/page/'+str(j)
		r = GetHTMLContent(url)
		html = etree.HTML(r.text)
		T1 = html.xpath('//*[@id="waterfall"]/div/a/div/span')	
		T2 = html.xpath('//*[@id="waterfall"]/div/a/div/span/date')	
		
		wb = load_workbook('data2.xlsx')
		ws = wb['Sheet1']

		for i in range(1,31):
			ws['A'+str(i+(j-1)*30)] = T1[i-1].text
			ws['B'+str(i+(j-1)*30)] = T2[i*2-2].text
			ws['C'+str(i+(j-1)*30)] = T2[i*2-1].text
		
		wb.save('data2.xlsx')

main()
